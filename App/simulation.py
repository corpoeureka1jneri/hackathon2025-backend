import pandas as pd
from .grid import Grid
from .models import Animal, Plant, Fungi

class Simulation:
    """Controla la ejecución de la simulación."""
    
    def __init__(self, grid: Grid, ticks: int):
        self.grid = grid  # Cuadrícula de 24x24
        self.ticks = ticks  # Número total de ticks a ejecutar
        self.current_tick = 0  # Contador de tiempo
        self.history = []  # Almacena estados de la simulación

    def run(self):
        """Ejecuta la simulación por el número de ticks configurado."""
        print(f"Iniciando simulación por {self.ticks} ticks...")

        # Guardar estado inicial
        self.save_state()

        for tick in range(1, self.ticks + 1):
            self.current_tick = tick
            self.update_grid()
            self.save_state()
        
        print("Simulación completada.")
        self.generate_report()

    def update_grid(self):
        """Actualiza todas las entidades y celdas en la cuadrícula."""
        self.grid.update()

    def save_state(self):
        """Guarda el estado actual de la cuadrícula."""
        state = []
        for x in range(self.grid.size):
            row = []
            for y in range(self.grid.size):
                tile = self.grid.tiles[x][y]
                cell_info = {
                    "tile_type": tile.type,
                    "entity": None,
                    "entity_state": None,
                    "entity_size": None,
                    "entity_hp": None
                }
                
                # Entidad principal en la celda
                if tile.entity:
                    cell_info["entity"] = type(tile.entity).__name__
                    cell_info["entity_state"] = tile.entity.state
                    
                    if isinstance(tile.entity, (Animal, Plant)):
                        cell_info["entity_size"] = tile.entity.size
                        
                    if isinstance(tile.entity, (Plant, Fungi)):
                        cell_info["entity_hp"] = tile.entity.hp
                
                # Entidades adicionales (hongos sobre entidades muertas)
                additional_entities = []
                for entity in self.grid.get_entities_at(x, y):
                    if entity != tile.entity:  # Evitar duplicados
                        entity_info = {
                            "type": type(entity).__name__,
                            "state": entity.state
                        }
                        if isinstance(entity, Fungi):
                            entity_info["hp"] = entity.hp
                        additional_entities.append(entity_info)
                
                if additional_entities:
                    cell_info["additional_entities"] = additional_entities
                
                row.append(cell_info)
            state.append(row)
        
        self.history.append(state)

    def generate_report(self):
        """Genera un archivo XLS con los resultados de la simulación, estilizado y con colores."""
        print("Generando reporte...")
        file_path = "output/simulation_results.xlsx"

        # Crear un Excel Writer con formato
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        
        # Definir estilos de celda para cada tipo
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Colores para los diferentes tipos
        styles = {
            # Tipos de suelo
            "water": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "rock": PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid"),
            "soil": PatternFill(start_color="C65911", end_color="C65911", fill_type="solid"),
            "soil+": PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
            
            # Entidades
            "animal-small": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
            "animal-big": PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid"),
            "plant-low": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            "plant-high": PatternFill(start_color="385723", end_color="385723", fill_type="solid"),
            "fungi": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
            
            # Estados
            "dead": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            
            # Estilo para encabezados
            "header": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        }
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Fuentes
        white_font = Font(color="FFFFFF", bold=True)
        black_font = Font(color="000000", bold=True)
        red_font = Font(color="FF0000", bold=True)
        
        # Alineación
        center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Para cada tick, crear una hoja con el estado
        for tick, state in enumerate(self.history):
            # Crear DataFrame base
            df = pd.DataFrame(index=range(self.grid.size), columns=range(self.grid.size))
            
            # Llenar datos y aplicar estilos
            for x in range(self.grid.size):
                for y in range(self.grid.size):
                    cell = state[x][y]
                    
                    # Siempre mostrar el tipo de tile como base
                    tile_type = cell["tile_type"]
                    
                    if cell["entity"]:
                        # Construir la información de la entidad
                        if cell["entity"] == "Animal":
                            entity_info = f"{cell['entity_size']}\n{cell['entity_state']}"
                        elif cell["entity"] == "Plant":
                            # Para plantas muertas, mostrar HP=0
                            hp_value = 0 if cell["entity_state"] == "dead" else cell["entity_hp"]
                            entity_info = f"{cell['entity_size']}\n{cell['entity_state']}\nHP:{hp_value}"
                        elif cell["entity"] == "Fungi":
                            entity_info = f"Fungi\n{cell['entity_state']}\nHP:{cell['entity_hp']}"
                        
                        # Combinar tipo de tile con información de la entidad
                        cell_value = f"[{tile_type}]\n{entity_info}"
                    else:
                        # Solo mostrar el tipo de tile
                        cell_value = f"[{tile_type}]"
                    
                    df.iloc[x, y] = cell_value

            # Guardar en Excel
            sheet_name = f'Tick {tick}'
            df.to_excel(writer, sheet_name=sheet_name)
            
            # Obtener la hoja actual
            worksheet = writer.sheets[sheet_name]
            
            # Ajustar ancho de columnas y alto de filas
            for column in worksheet.columns:
                worksheet.column_dimensions[column[0].column_letter].width = 18
            for row in worksheet.rows:
                worksheet.row_dimensions[row[0].row].height = 60  # Aumentar altura para más información

            # Aplicar estilos a las celdas
            for x in range(self.grid.size):
                for y in range(self.grid.size):
                    cell = state[x][y]
                    excel_cell = worksheet.cell(row=x+2, column=y+2)  # +2 por el offset del índice
                    
                    # Aplicar estilo según el contenido
                    if cell["entity"]:
                        if cell["entity"] == "Animal":
                            style = styles[cell["entity_size"]]
                        elif cell["entity"] == "Plant":
                            style = styles[cell["entity_size"]]
                        elif cell["entity"] == "Fungi":
                            style = styles["fungi"]
                            
                        if cell["entity_state"] == "dead":
                            style = styles["dead"]
                    else:
                        style = styles[cell["tile_type"]]
                    
                    excel_cell.fill = style
                    excel_cell.border = thin_border
                    excel_cell.alignment = center_aligned
                    
                    # Ajustar color de fuente según el fondo
                    if cell["tile_type"] in ["water", "rock"] or (cell["entity"] and cell["entity_state"] == "dead"):
                        excel_cell.font = white_font
                    else:
                        excel_cell.font = black_font

            # Estilizar encabezados
            for cell in worksheet[1]:
                cell.fill = styles["header"]
                cell.font = white_font
                cell.alignment = center_aligned
                cell.border = thin_border
            
            for cell in worksheet['A']:
                cell.fill = styles["header"]
                cell.font = white_font
                cell.alignment = center_aligned
                cell.border = thin_border

        # Crear hoja de resumen
        summary_data = self.generate_summary()
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Resumen')
        
        # Estilizar hoja de resumen
        worksheet = writer.sheets['Resumen']
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = max_length + 2

        # Aplicar estilos al resumen
        for row in worksheet.rows:
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_aligned
                if cell.row == 1:  # Encabezados
                    cell.fill = styles["header"]
                    cell.font = white_font

        writer.close()
        print(f"Reporte guardado en {file_path}")

    def generate_summary(self):
        """Genera un resumen de la simulación para incluir en el reporte."""
        summary = []
        
        for tick, state in enumerate(self.history):
            # Contar entidades por tipo
            counts = {
                "tick": tick,
                "animal_small": 0,
                "animal_big": 0,
                "plant_low": 0,
                "plant_high": 0,
                "fungi": 0,
                "dead_entities": 0,
                "water_tiles": 0,
                "rock_tiles": 0,
                "soil_tiles": 0,
                "soil+_tiles": 0
            }
            
            for x in range(self.grid.size):
                for y in range(self.grid.size):
                    cell = state[x][y]
                    
                    # Contar tipos de suelo
                    counts[f"{cell['tile_type']}_tiles"] += 1
                    
                    # Contar entidades
                    if cell["entity"] == "Animal":
                        if cell["entity_size"] == "animal-small":
                            counts["animal_small"] += 1
                        else:
                            counts["animal_big"] += 1
                    elif cell["entity"] == "Plant":
                        if cell["entity_size"] == "plant-low":
                            counts["plant_low"] += 1
                        else:
                            counts["plant_high"] += 1
                    elif cell["entity"] == "Fungi":
                        counts["fungi"] += 1
                        
                    if cell["entity_state"] == "dead":
                        counts["dead_entities"] += 1
            
            summary.append(counts)
        
        return summary
